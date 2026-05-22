#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

import clickhouse_connect
import yaml
from openpyxl import load_workbook


@dataclass
class Config:
    clickhouse: dict[str, Any]
    archive_dir: str | None
    delete_after_load: bool
    min_file_age_seconds: int


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Load company registry xlsx into ClickHouse")
    p.add_argument("--config", required=True)
    p.add_argument("--landing", required=True)
    return p.parse_args()


def load_config(path: str) -> Config:
    raw = yaml.safe_load(Path(path).read_text(encoding="utf-8"))
    return Config(
        clickhouse=raw["clickhouse"],
        archive_dir=raw.get("archive_dir"),
        delete_after_load=bool(raw.get("delete_after_load", False)),
        min_file_age_seconds=int(raw.get("min_file_age_seconds", 180)),
    )


def ch_client(conf: Config):
    return clickhouse_connect.get_client(
        host=conf.clickhouse["host"],
        port=conf.clickhouse.get("port", 8123),
        username=conf.clickhouse.get("username", "default"),
        password=conf.clickhouse.get("password", ""),
        database=conf.clickhouse.get("database", "analytics_1c"),
    )


def normalize_company_key(value: str) -> str:
    text = (value or "").upper().replace("Ё", "Е")
    text = re.sub(r"(^|\s)20\d{2}($|\s)", " ", text)
    text = re.sub(r"[^0-9A-ZА-Я]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def archive_or_delete(conf: Config, path: Path) -> None:
    if conf.archive_dir:
        archive_root = Path(conf.archive_dir) / "registry"
        archive_root.mkdir(parents=True, exist_ok=True)
        shutil.move(str(path), archive_root / path.name)
        return
    if conf.delete_after_load:
        path.unlink(missing_ok=True)


def parse_registry(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    now = datetime.now(UTC).replace(tzinfo=None, microsecond=0)

    tax_map: dict[str, tuple[str, str]] = {}
    if "Лист2" in wb.sheetnames:
        ws = wb["Лист2"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            company_name = as_text(row[1] if len(row) > 1 else "")
            if not company_name:
                continue
            tax_map[normalize_company_key(company_name)] = (
                as_text(row[2] if len(row) > 2 else ""),
                as_text(row[3] if len(row) > 3 else ""),
            )

    rows: list[dict[str, Any]] = []
    if "ОСНОВНОЙ" not in wb.sheetnames:
        return rows

    ws = wb["ОСНОВНОЙ"]
    top_headers = [as_text(v) for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    manager_headers = [as_text(v) for v in next(ws.iter_rows(min_row=2, max_row=2, values_only=True))]

    col_specs: list[dict[str, Any]] = []
    current_manager: dict[str, Any] | None = None
    for idx in range(1, len(manager_headers)):
        manager = manager_headers[idx]
        top = top_headers[idx] if idx < len(top_headers) else ""
        if manager:
            current_manager = {
                "col": idx,
                "assignee_name": manager,
                "meta_col": None,
                "meta_label": "",
                "registry_status": "active",
            }
            col_specs.append(current_manager)
            continue
        if top and "исключ" in top.lower():
            current_manager = {
                "col": idx,
                "assignee_name": "",
                "meta_col": None,
                "meta_label": "",
                "registry_status": "excluded",
            }
            col_specs.append(current_manager)
            continue
        if current_manager is not None:
            current_manager["meta_col"] = idx
            current_manager["meta_label"] = top or "meta"
            current_manager = None

    for row in ws.iter_rows(min_row=3, values_only=True):
        for spec in col_specs:
            company_name = as_text(row[spec["col"]] if spec["col"] < len(row) else "")
            if not company_name:
                continue
            meta_value = ""
            if spec.get("meta_col") is not None:
                meta_value = as_text(row[spec["meta_col"]] if spec["meta_col"] < len(row) else "")
            company_key = normalize_company_key(company_name)
            inn, kpp = tax_map.get(company_key, ("", ""))
            rows.append(
                {
                    "ts": now,
                    "source_file": path.name,
                    "source_sheet": "ОСНОВНОЙ",
                    "company_name": company_name,
                    "company_key": company_key,
                    "assignee_name": spec["assignee_name"],
                    "registry_status": spec["registry_status"],
                    "share_text": meta_value if "ключ" not in spec.get("meta_label", "").lower() else "",
                    "key_contour": 1 if meta_value.upper() == "ЕСТЬ" and "ключ" in spec.get("meta_label", "").lower() else 0,
                    "inn": inn,
                    "kpp": kpp,
                }
            )

    return rows


def main() -> int:
    args = parse_args()
    conf = load_config(args.config)
    client = ch_client(conf)
    landing = Path(args.landing)
    if not landing.exists():
        return 0

    for path in sorted(p for p in landing.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"):
        age_seconds = max(0, int((datetime.now(UTC) - datetime.fromtimestamp(path.stat().st_mtime, UTC)).total_seconds()))
        if age_seconds < conf.min_file_age_seconds:
            print(f"skip registry: {path.name} age={age_seconds}s < min_file_age_seconds={conf.min_file_age_seconds}")
            continue
        rows = parse_registry(path)
        if rows:
            client.insert(
                "analytics_1c.raw_1c_company_registry",
                [[row["source_file"], row["source_sheet"], json.dumps({k: (v.isoformat() if isinstance(v, datetime) else v) for k, v in row.items()}, ensure_ascii=False)] for row in rows],
                column_names=["source_file", "source_sheet", "payload"],
            )
            client.insert(
                "analytics_1c.company_registry",
                [[row["ts"], row["source_file"], row["source_sheet"], row["company_name"], row["company_key"], row["assignee_name"], row["registry_status"], row["share_text"], row["key_contour"], row["inn"], row["kpp"]] for row in rows],
                column_names=["ts", "source_file", "source_sheet", "company_name", "company_key", "assignee_name", "registry_status", "share_text", "key_contour", "inn", "kpp"],
            )
            print(f"loaded registry: {path.name} rows={len(rows)}")
        archive_or_delete(conf, path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
